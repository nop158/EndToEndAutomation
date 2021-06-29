import openpyxl

def data_generate():
    wk = openpyxl.load_workbook("D:\selenium.xlsx")
    sh = wk['Sheet1']
    r =  sh.max_row
    li = []
    li1 = []
    for i in range(1,r + 1):
        li1 = []
        un = sh.cell(i, 1)
        pw = sh.cell(i, 2)
        li1.insert(0, un.value)
        li1.insert(1, pw.value)
        li.insert(i-1, li1)
    return li